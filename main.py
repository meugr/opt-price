#!/usr/bin/python3
from config import Config
import requests
import json
import os
import openpyxl
from datetime import datetime as dt
from shutil import copyfile
import smtplib
try:
    import lxml.etree.ElementTree as etree
except ImportError:
    import xml.etree.ElementTree as etree


def get_xml_file(xml_link, xml_name):
    """
    Get and write XML from URL.
    """
    xml_price = requests.get(xml_link).content
    with open(xml_name, "wb") as f:
        f.write(xml_price)


def make_dict(xml_name, urls_name):
    """
    Сreates a dictionary from an XML file accessible by xml_name.
    """
    dict_values = {}
    with open(urls_name, "r") as f:
        urls = json.load(f)  # urls = {art1: url1, art2: url2, ...}
    tree = etree.parse(xml_name)
    root = tree.getroot()
    categories = root[0][4]
    offers = root[0][5]
    for offer in offers:
        identificator = offer.get('id')
        available = offer.get('available')
        name = offer.find('name').text
        price = offer.find('price').text
        weight = offer.find('param[@name="Вес"]').text

        category_id = offer.find('categoryId').text
        category = categories.find('category[@id="{}"]'.format(category_id)).text

        vendor = offer.find('vendor')  # maybe None
        vendor = vendor.text if vendor is not None else 'Без бренда'

        flavor = offer.find('param[@name="Вкус"]')  # maybe None
        flavor = flavor.text if flavor is not None else 'без вкусов'

        url = urls.get(identificator) if identificator in urls.keys() else ''

        dict_values[identificator] = {'name': name,
                                      'price': price,
                                      'weight': weight,
                                      'category': category,
                                      'vendor': vendor,
                                      'flavor': flavor,
                                      'available': available,
                                      'url': url}
    return dict_values


def write_in_excel(dict_price, excel_name, current_date):
    """
    Writes values from dictionary to an Excel price list.
    Clear 50 rows after last row.
    """
    row = 7  # starting row for write
    wb = openpyxl.load_workbook(excel_name)
    ws = wb.active
    current_date = dt.now().strftime("%d.%m.%Y")

    for art in dict_price:
        row += 1
        ws['A{}'.format(row)] = art
        ws['B{}'.format(row)] = dict_price.get(art).get('name')
        ws['C{}'.format(row)] = dict_price.get(art).get('category')
        ws['D{}'.format(row)] = dict_price.get(art).get('vendor')
        ws['E{}'.format(row)] = dict_price.get(art).get('flavor')
        ws['F{}'.format(row)] = dict_price.get(art).get('weight')
        ws['G{}'.format(row)] = dict_price.get(art).get('price')
        ws['H{}'.format(row)] = dict_price.get(art).get('url')
    ws['A1'] = 'Цены и наличие\nактуальны на\n{}'.format(current_date)
    for i in range(row + 1, row + 50):  # clear 50 row after last row
        for j in range(1, 9):
            ws.cell(row=i, column=j, value='')
    wb.save(excel_name)


def send_error_message(email_logpass_sender, addressee, e):
    """
    Send e-mail with error message.
    email_logpass_sender = (login, password)
    """
    HOST = ('smtp.yandex.ru', 465)
    SUBJECT = 'Error opt-price generator'
    TO = addressee
    FROM = email_logpass_sender[0]
    text = 'Error generating xlsx, urgently check the operation \
of the script "opt-price".\n\n{}'.format(e)

    BODY = "\r\n".join((
           "From: %s" % FROM,
           "To: %s" % TO,
           "Subject: %s" % SUBJECT,
           "", text))

    server = smtplib.SMTP_SSL(*HOST)
    server.login(*email_logpass_sender)
    server.sendmail(FROM, TO, BODY)
    server.quit()


try:
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    current_date = dt.now().strftime("%d.%m.%Y")

    get_xml_file(Config.xml_link, Config.xml_name)
    dict_price = make_dict(Config.xml_name, Config.urls_name)
    write_in_excel(dict_price, Config.excel_name, current_date)
    try:
        os.mkdir('backup')  # creates a directory when you first start
    except FileExistsError:
        print('Directory "backup" already exist')
    backup_path = os.path.join('backup', '{}.xls'.format(current_date))
    copyfile(Config.excel_name, backup_path)
except BaseException as e:
    print('ERROR!!\n', e)
    send_error_message(Config.email_logpass_sender, Config.addressee, e)
